import openpyxl
import sys

path_excel='C:/_BangNguyen/documents_bosch/Workspace/_Scrip_check_output_ASW/ok/TD_CRB_Pressure2Volume_MT_097.xlsm'
wb = openpyxl.load_workbook(path_excel, data_only=True)
sheet_Summary = wb['Summary']
print ("                C0 C1........................", sheet_Summary['A6'].value)
print ("                MCDC.........................", sheet_Summary['B6'].value)
print ("    Reason : ", sheet_Summary['C6'].value)
print (" ======================== ")

try:
    sheet_Testcases = wb['Testcases']
except Exception as e:  
    sheet_Testcases = wb.worksheets[3]

'''Check input'''
TC_No = 1
Row_Tolerance = 18
Row_Type = 19
Row_Max = 20
Row_Min = 21
Row_Title = 22
Row_Name_Var = 23

max_row_table = sheet_Testcases.max_row
max_column_table = sheet_Testcases.max_column
print (max_row_table)
for i in range(Row_Name_Var,max_row_table):
    if sheet_Testcases.cell(i, TC_No).value == 'None':
        print (i)
        max_row_table = i + 23
        break
        
print (max_row_table)
'''Check TM name'''
if sheet_Testcases.cell(Row_Name_Var, TC_No).value == 'None':
    print ("Miss TM_ name")


'''Check input cont'''
for i in range(1,max_column_table+1):
    if str(sheet_Testcases.cell(Row_Title, i).value) == 'INPUTS':
        loc_input = i
        break

for i in range(loc_input + 1,max_column_table+1):
    if str(sheet_Testcases.cell(Row_Title, i).value) != 'None':
        loc_end_input = i
        break
 
for i in range(loc_input,loc_end_input):
    if str(sheet_Testcases.cell(Row_Type, i).value) == 'cont':
        value_tol=float(sheet_Testcases.cell(Row_Tolerance, i).value)
        value_max=float(sheet_Testcases.cell(Row_Max, i).value)
        value_min=float(sheet_Testcases.cell(Row_Min, i).value)

        flag_not_out_max=0
        flag_not_out_min=0
        flag_not_mid_value=0
        '''Check input out max'''
        for j in range(24,max_row_table+1): 
            if float(sheet_Testcases.cell(j, i).value) > value_max :
                flag_not_out_max = 0
                break
            else :
                flag_not_out_max = 1
        '''Check input out min'''
        for j in range(24,max_row_table+1): 
            print (sheet_Testcases.cell(j, i).value)
            if float(sheet_Testcases.cell(j, i).value) < value_min :
                flag_not_out_min = 0
                break
            else :
                flag_not_out_min = 1        
        '''Check input not mid value'''
        for j in range(24,max_row_table+1): 
            if float(sheet_Testcases.cell(j, i).value) < value_max and float(sheet_Testcases.cell(j, i).value) > value_min :
                flag_not_mid_value = 0 
                break
            else :
                flag_not_mid_value = 1   
        '''print resuit'''                
        if flag_not_out_max == 1:
            print ("Not out max: ",str(sheet_Testcases.cell(Row_Name_Var, i).value))
        if flag_not_out_min == 1:
            print ("Not out min: ",str(sheet_Testcases.cell(Row_Name_Var, i).value))
        if flag_not_mid_value == 1:
            print ("Not out mid value: ",str(sheet_Testcases.cell(Row_Name_Var, i).value))

'''Check input log'''
for i in range(loc_input,loc_end_input):
    if str(sheet_Testcases.cell(Row_Type, i).value) == 'log':
        flag_true = 0   
        flag_false = 0     
        for j in range(24,max_row_table+1):
            if str(sheet_Testcases.cell(j, i).value) == 'True':
                flag_true = 1
            if str(sheet_Testcases.cell(j, i).value) == 'False':
                flag_false = 1
        if flag_true == 1 and flag_false == 1:
            pass
        else :
            print ("Thieu TRUE/FALSE: ",str(sheet_Testcases.cell(Row_Name_Var, i).value))
            
'''Check input enum'''
for i in range(loc_input,loc_end_input):
    if str(sheet_Testcases.cell(Row_Type, i).value) == 'enum':
        print ("Check enum: ", str(sheet_Testcases.cell(Row_Name_Var, i).value))


wb.close()